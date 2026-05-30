from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.ingestion.types import ParsedBlock


def _stringify(value: object) -> str:
    if value is None:
        return ""
    # Collapse internal newlines/tabs so table-aware chunking can safely use
    # line boundaries to represent worksheet rows.
    return " ".join(str(value).split())


class XlsxParser:
    def __init__(self, rows_per_block: int = 20) -> None:
        self.rows_per_block = rows_per_block

    def parse(self, file_path: Path) -> list[ParsedBlock]:
        workbook = load_workbook(file_path, read_only=False, data_only=True)
        blocks: list[ParsedBlock] = []
        try:
            for worksheet in workbook.worksheets:
                merged_value_map, merged_group_columns = self._build_merged_metadata(worksheet)
                row_records = [
                    (
                        row_index,
                        [
                            self._resolve_cell_value(
                                row_index=row_index,
                                column_index=column_index,
                                raw_value=cell.value,
                                merged_value_map=merged_value_map,
                            )
                            for column_index, cell in enumerate(row, start=1)
                        ],
                    )
                    for row_index, row in enumerate(worksheet.iter_rows(), start=1)
                ]
                non_empty_records = [(row_number, row) for row_number, row in row_records if any(cell for cell in row)]
                if not non_empty_records:
                    continue

                headers = [
                    header if header else f"column_{index + 1}"
                    for index, header in enumerate(non_empty_records[0][1])
                ]
                data_records = non_empty_records[1:] or [(non_empty_records[0][0], headers)]
                header_line = " | ".join(headers)
                inferred_group_columns = self._infer_group_columns([row for _, row in data_records])
                group_columns = sorted(merged_group_columns or inferred_group_columns)

                current_batch: list[str] = []
                current_row_numbers: list[int] = []
                current_row_records: list[dict[str, object]] = []
                current_start = data_records[0][0] if data_records else non_empty_records[0][0]
                current_end = current_start - 1
                inherited_prefix_values = [""] * len(headers)

                for offset, raw_row in data_records:
                    effective_row = self._apply_prefix_inheritance(raw_row, inherited_prefix_values)
                    entries = []
                    structured_fields: dict[str, str] = {}
                    similar_questions: list[str] = []
                    for index, cell_value in enumerate(effective_row):
                        if not cell_value:
                            continue
                        header = headers[index] if index < len(headers) else f"column_{index + 1}"
                        entries.append(f"{header}: {cell_value}")
                        if header.startswith("相似问题"):
                            similar_questions.append(cell_value)
                        else:
                            structured_fields[header] = cell_value
                    if not entries:
                        continue
                    row_text = "; ".join(entries)
                    current_batch.append(row_text)
                    current_row_numbers.append(offset)
                    current_row_records.append(
                        {
                            "row_number": offset,
                            "text": row_text,
                            "group_key": [effective_row[index] for index in group_columns if index < len(effective_row) and effective_row[index]],
                            "fields": structured_fields,
                            "similar_questions": similar_questions,
                        }
                    )
                    current_end = offset

                    if len(current_batch) >= self.rows_per_block:
                        blocks.append(
                            ParsedBlock(
                                text="\n".join([header_line, *current_batch]),
                                chunk_type="table",
                                sheet_name=worksheet.title,
                                row_start=current_start,
                                row_end=current_end,
                                metadata={
                                    "parser": "openpyxl",
                                    "headers": headers,
                                    "_row_numbers": list(current_row_numbers),
                                    "_table_rows": list(current_row_records),
                                    "_group_columns": list(group_columns),
                                },
                            )
                        )
                        current_batch = []
                        current_row_numbers = []
                        current_row_records = []
                        current_start = offset + 1

                if current_batch:
                    blocks.append(
                        ParsedBlock(
                            text="\n".join([header_line, *current_batch]),
                            chunk_type="table",
                            sheet_name=worksheet.title,
                            row_start=current_start,
                            row_end=current_end,
                            metadata={
                                "parser": "openpyxl",
                                "headers": headers,
                                "_row_numbers": list(current_row_numbers),
                                "_table_rows": list(current_row_records),
                                "_group_columns": list(group_columns),
                            },
                        )
                    )
        finally:
            workbook.close()

        return blocks

    @staticmethod
    def _apply_prefix_inheritance(row: list[str], inherited_prefix_values: list[str]) -> list[str]:
        effective_row = list(row)
        first_non_empty_index = next((index for index, value in enumerate(effective_row) if value), None)
        if first_non_empty_index is not None:
            for index in range(first_non_empty_index):
                if not effective_row[index] and inherited_prefix_values[index]:
                    effective_row[index] = inherited_prefix_values[index]

        for index, value in enumerate(effective_row):
            if value:
                inherited_prefix_values[index] = value

        return effective_row

    @staticmethod
    def _resolve_cell_value(
        *,
        row_index: int,
        column_index: int,
        raw_value: object,
        merged_value_map: dict[tuple[int, int], str],
    ) -> str:
        if raw_value is not None:
            return _stringify(raw_value)
        return merged_value_map.get((row_index, column_index), "")

    @staticmethod
    def _build_merged_metadata(worksheet) -> tuple[dict[tuple[int, int], str], set[int]]:
        merged_value_map: dict[tuple[int, int], str] = {}
        merged_group_columns: set[int] = set()
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.max_row <= merged_range.min_row:
                continue
            anchor_value = _stringify(worksheet.cell(merged_range.min_row, merged_range.min_col).value)
            for row_index in range(merged_range.min_row, merged_range.max_row + 1):
                for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_value_map[(row_index, column_index)] = anchor_value
            merged_group_columns.add(merged_range.min_col - 1)
        return merged_value_map, merged_group_columns

    @staticmethod
    def _infer_group_columns(data_rows: list[list[str]]) -> set[int]:
        if not data_rows:
            return set()

        total_rows = len(data_rows)
        inferred_columns: set[int] = set()
        for column_index in range(len(data_rows[0])):
            explicit_count = sum(1 for row in data_rows if column_index < len(row) and row[column_index])
            fill_ratio = explicit_count / total_rows if total_rows else 0
            if fill_ratio >= 0.8:
                break
            inferred_columns.add(column_index)

        return inferred_columns
